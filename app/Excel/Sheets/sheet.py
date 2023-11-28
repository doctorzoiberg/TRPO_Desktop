from .Styles import styles
from openpyxl.utils.cell import get_column_letter


def create_sheet(work_book, data):
    platform = data[0][0]

    work_sheet = work_book.active

    work_sheet.title = platform

    return work_sheet


def set_size(work_sheet, data):
    column_name_list = data[1]
    number_columns = len(column_name_list)
    column_numbers = range(number_columns)

    for column_number in column_numbers:
        column_width = 0

        for row_list in data[1:]:
            value_cell_len = len(row_list[column_number])
            if value_cell_len > column_width:
                column_width = value_cell_len

        column_letter = get_column_letter(column_number + 1)
        work_sheet.column_dimensions[column_letter].width = column_width + 10

    work_sheet.row_dimensions[1].height = 30
    work_sheet.row_dimensions[2].height = 20


def set_titles(work_sheet, data):
    platform = data[0][0]
    work_sheet['A1'] = platform

    column_name_list = data[1]
    number_columns = len(column_name_list)
    last_column_letter = get_column_letter(number_columns)
    work_sheet.merge_cells(f'A1:{last_column_letter}1')

    work_sheet.append(column_name_list)


def set_data(work_sheet, data):
    for row_list in data[2:]:
        work_sheet.append(row_list)


def set_styles(work_sheet, data):
    number_rows = len(data)
    column_name_list = data[1]
    number_columns = len(column_name_list)
    row_numbers = range(2, (number_rows + 1))
    column_numbers = range(1, (number_columns + 1))

    align = styles.alignment()
    work_sheet['A1'].alignment = align
    for row_number in row_numbers:
        for column_number in column_numbers:
            work_sheet.cell(row=row_number, column=column_number).alignment = align

    title_font = styles.title_font()
    work_sheet['A1'].font = title_font
    for column_number in column_numbers:
        work_sheet.cell(row=2, column=column_number).font = styles.title_font()

    border = styles.border('bottom')
    for column_number in column_numbers:
        work_sheet.cell(row=1, column=column_number).border += border
        work_sheet.cell(row=2, column=column_number).border += border
        work_sheet.cell(row=number_rows, column=column_number).border += border

    border = styles.border('right')
    work_sheet.cell(row=1, column=number_columns).border += border
    for column_number in column_numbers:
        work_sheet.cell(row=2, column=column_number).border += border
    for row_number in row_numbers:
        work_sheet.cell(row=row_number, column=number_columns).border += border
