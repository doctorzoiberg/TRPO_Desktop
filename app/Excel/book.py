from openpyxl import Workbook as book
from .Sheets import sheet


def create_book():
    return book()


def save_book(work_book, path):
    work_book.save(path)


def create_sheet(work_book, data):
    work_sheet = sheet.create_sheet(work_book, data)
    sheet.set_size(work_sheet, data)
    sheet.set_titles(work_sheet, data)
    sheet.set_data(work_sheet, data)
    sheet.set_styles(work_sheet, data)


def main(data, path):
    work_book = create_book()
    create_sheet(work_book, data)
    save_book(work_book, path)
