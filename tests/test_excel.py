import pytest
import openpyxl
from app.Excel import book
from app.Parser import parser
import pathlib


url_android = 'https://www.antutu.com/en/ranking/rank1.htm'
url_ios = 'https://www.antutu.com/en/ranking/ios1.htm'
url_ai = 'https://www.antutu.com/en/ranking/ai1.htm'


@pytest.fixture(name='work_book')
def create_book():
    return openpyxl.Workbook()


def get_an_data():
    return parser.get_phones_data(parser.parse_text(url_android))

def get_io_data():
    return parser.get_phones_data(parser.parse_text(url_android))

def get_ai_data():
    return parser.get_phones_data(parser.parse_text(url_android))


def test_create_book():
    bk = book.create_book()
    assert type(bk) == openpyxl.workbook.workbook.Workbook


def test_create_phone_sheet(work_book):
    book.create_phone_sheet(work_book, 'Android', get_an_data())
    assert work_book.sheetnames[0] == 'Android'
    book.create_phone_sheet(work_book, 'IOS', get_io_data())
    assert work_book.sheetnames[0] == 'IOS'
    book.create_phone_sheet(work_book, 'Ai', get_ai_data())
    assert work_book.sheetnames[0] == 'Ai'


def test_save_book(work_book, tmp_path):
    path = pathlib.Path(tmp_path, 'test.xlsx')
    book.save_book(work_book, path)
    assert path.exists()


def test_main(tmp_path):
    path = pathlib.Path(tmp_path, 'test_main.xlsx')
    book.main(get_an_data(), 'Android', path)
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == 'Android'
    sheet = wb.active
    assert sheet['A1'].value == 'Android'
